"""
生成详细分析 Excel：
  Sheet1 - 相关性矩阵（11策略 + 融合组合，含颜色热力）
  Sheet2 - 滚动12个月分析（11策略 + 融合组合）
  Sheet3 - 月度收益明细表（所有策略和组合，年×月）
"""
import sys, math, datetime
sys.path.insert(0, '.')

import heatmap_analysis
import correlation_matrix as cm_mod
import rolling_analysis as ra_mod
from analysis_utils import read_monthly_pnl

from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as xl_numbers)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles import Color

DEFAULT_FILES = {
    'BTC_ema': '/Users/huminghe/Downloads/strategy_ema_btc_OKX_BTCUSDT.P_2026-05-18_bccdc.xlsx',
    'BTC_v2':  '/Users/huminghe/Downloads/v2_strategy_btc_OKX_BTCUSDT.P_2026-05-18_89f2b.xlsx',
    'BTC_v3':  '/Users/huminghe/Downloads/v3_205m_strategy_btc_OKX_BTCUSDT.P_2026-05-18_7f0bc.xlsx',
    'ETH_ema': '/Users/huminghe/Downloads/strategy_ema_eth_OKX_ETHUSDT.P_2026-05-18_6d950.xlsx',
    'ETH_v2':  '/Users/huminghe/Downloads/v2_strategy_eth_OKX_ETHUSDT.P_2026-05-18_7eba3.xlsx',
    'ETH_v3':  '/Users/huminghe/Downloads/v3_3h_strategy_eth_OKX_ETHUSDT.P_2026-05-18_f84bb.xlsx',
    'SOL_ema': '/Users/huminghe/Downloads/strategy_ema_sol_OKX_SOLUSDT.P_2026-05-18_f737f.xlsx',
    'SOL_v2':  '/Users/huminghe/Downloads/v2_strategy_sol_OKX_SOLUSDT.P_2026-05-18_2c3a0.xlsx',
    'SOL_v3':  '/Users/huminghe/Downloads/v3_3h_strategy_sol_OKX_SOLUSDT.P_2026-05-18_a51fb.xlsx',
    'DOGE_v2': '/Users/huminghe/Downloads/v2_strategy_doge_OKX_DOGEUSDT.P_2026-05-18_2b9fa.xlsx',
    'DOGE_v3': '/Users/huminghe/Downloads/v3_205m_strategy_doge_OKX_DOGEUSDT.P_2026-05-18_283df.xlsx',
}
FILES = DEFAULT_FILES  # 向后兼容

INITIAL = 50_000.0

# ─── 样式 ─────────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill('solid', start_color='1F4E79')
HDR_FONT  = Font(bold=True, color='FFFFFF', name='Arial', size=10)
GRP_FILLS = {
    'BTC':  PatternFill('solid', start_color='FFF2CC'),
    'ETH':  PatternFill('solid', start_color='DAE8FC'),
    'SOL':  PatternFill('solid', start_color='E1D5E7'),
    'DOGE': PatternFill('solid', start_color='D5E8D4'),
    'combo':PatternFill('solid', start_color='F8CECC'),
}
THIN  = Side(style='thin',   color='BFBFBF')
MED   = Side(style='medium', color='888888')
BORD  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CTR   = Alignment(horizontal='center', vertical='center')
LEFT  = Alignment(horizontal='left',   vertical='center')
BODY  = Font(name='Arial', size=10)
BOLD  = Font(bold=True, name='Arial', size=10)

def asset_of(key):
    if '_' in key:
        return key.split('_')[0]
    for a in ['BTC','ETH','SOL','DOGE']:
        if a in key:
            return a
    return 'combo'

def fill_of(key):
    return GRP_FILLS.get(asset_of(key), GRP_FILLS['combo'])

def hdr(ws, row, col, val, width=None):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = CTR; c.border = BORD
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c

def cell(ws, row, col, val, fmt=None, bold=False, fill=None, align=CTR):
    c = ws.cell(row=row, column=col, value=val)
    c.font = BOLD if bold else BODY
    c.alignment = align
    c.border = BORD
    if fmt:  c.number_format = fmt
    if fill: c.fill = fill
    return c

def run(files: dict[str, str], base: str, out_dir: str, date_str: str = None):
    """
    files: {key: 绝对路径}，base 传空字符串
    out_dir: 输出目录
    date_str: 日期字符串，用于文件名（默认今天）
    """
    import os, datetime as _dt
    if date_str is None:
        date_str = _dt.date.today().isoformat()

    # ─── 读取数据 ─────────────────────────────────────────────────────────────
    print('读取数据...')
    combos_monthly = heatmap_analysis.build_combo_monthly(files, base=base)

    available_combos = {k: v for k, v in combos_monthly.items()
                        if k not in files and v}
    print(f'  可用融合组合: {list(available_combos.keys())}')

    all_monthly: dict[str, dict] = {}
    for k, fp in files.items():
        all_monthly[k] = read_monthly_pnl(base + fp)
    all_monthly.update(available_combos)

    ALL_KEYS = list(files.keys()) + list(available_combos.keys())

    def pearson(a, b):
        n = len(a)
        if n < 2: return float('nan')
        ma, mb = sum(a)/n, sum(b)/n
        cov = sum((a[i]-ma)*(b[i]-mb) for i in range(n))
        sa = math.sqrt(sum((x-ma)**2 for x in a))
        sb = math.sqrt(sum((x-mb)**2 for x in b))
        return cov/(sa*sb) if sa > 1e-9 and sb > 1e-9 else float('nan')

    def aligned_series(k1, k2):
        common = sorted(set(all_monthly[k1]) & set(all_monthly[k2]))
        if len(common) < 6:
            common = sorted(set(all_monthly[k1]) | set(all_monthly[k2]))
        a = [all_monthly[k1].get(m, 0.0) for m in common]
        b = [all_monthly[k2].get(m, 0.0) for m in common]
        return a, b, len(common)

    # ─── 滚动指标（含融合组合）────────────────────────────────────────────────
    def rolling_metrics_from_monthly(monthly, window=12):
        months = sorted(monthly.keys())
        results = []
        for i in range(window-1, len(months)):
            wm = months[i-window+1:i+1]
            pnls = [monthly.get(m, 0.0) for m in wm]
            returns = [p/INITIAL for p in pnls]
            mean_r = sum(returns)/len(returns)
            std_r  = math.sqrt(sum((r-mean_r)**2 for r in returns)/len(returns))
            sharpe = mean_r/std_r*math.sqrt(12) if std_r > 1e-9 else 0.0
            cum, peak, max_dd = 0.0, 0.0, 0.0
            for p in pnls:
                cum += p
                if cum > peak: peak = cum
                dd = (peak-cum)/INITIAL*100
                if dd > max_dd: max_dd = dd
            win_rate = sum(1 for p in pnls if p > 0)/len(pnls)*100
            results.append((datetime.date(wm[-1][0], wm[-1][1], 1),
                            sharpe, max_dd, win_rate))
        return results

    rolling_data = {}
    for k in ALL_KEYS:
        m = all_monthly.get(k, {})
        if m:
            rolling_data[k] = rolling_metrics_from_monthly(m, window=12)

    print(f'  滚动数据: {len(rolling_data)} 条')

    # ─── 创建 Excel ───────────────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 1: 相关性矩阵
    # ═══════════════════════════════════════════════════════════════════════════
    ws1 = wb.create_sheet('相关性矩阵')

    # 标题
    ws1.merge_cells(f'A1:{get_column_letter(len(ALL_KEYS)+2)}1')
    ws1['A1'] = f'策略月度收益相关性矩阵（{len(ALL_KEYS)} 个策略/组合）'
    ws1['A1'].font = Font(bold=True, name='Arial', size=12, color='1F4E79')
    ws1['A1'].alignment = CTR
    ws1['A1'].fill = PatternFill('solid', start_color='D6E4F0')

    # 列标题行
    hdr(ws1, 2, 1, '策略/组合', width=14)
    hdr(ws1, 2, 2, '类型', width=8)
    for j, k in enumerate(ALL_KEYS, 3):
        c = ws1.cell(row=2, column=j, value=k)
        c.fill = fill_of(k); c.font = BOLD
        c.alignment = CTR; c.border = BORD
        ws1.column_dimensions[get_column_letter(j)].width = 9

    # 数据行
    for i, ki in enumerate(ALL_KEYS):
        row = i + 3
        # 行标题
        c = ws1.cell(row=row, column=1, value=ki)
        c.fill = fill_of(ki); c.font = BOLD; c.alignment = LEFT; c.border = BORD
        # 类型
        typ = '单策略' if '_' in ki else '融合'
        cell(ws1, row, 2, typ, fill=fill_of(ki))
        # 相关系数
        for j, kj in enumerate(ALL_KEYS, 3):
            a, b, n = aligned_series(ki, kj)
            r = pearson(a, b)
            c = ws1.cell(row=row, column=j)
            if math.isnan(r):
                c.value = '-'
            else:
                c.value = round(r, 3)
                c.number_format = '0.000'
            c.alignment = CTR; c.border = BORD; c.font = BODY
            # 对角线
            if ki == kj:
                c.fill = PatternFill('solid', start_color='D9D9D9')
                c.font = Font(bold=True, name='Arial', size=10)

    # 条件格式：相关系数热力（跳过对角线，用 ColorScale）
    data_range = f'C3:{get_column_letter(len(ALL_KEYS)+2)}{len(ALL_KEYS)+2}'
    ws1.conditional_formatting.add(data_range, ColorScaleRule(
        start_type='num', start_value=-1, start_color='D73027',
        mid_type='num',   mid_value=0,    mid_color='FFFFBF',
        end_type='num',   end_value=1,    end_color='4575B4',
    ))

    # 附表：高相关对（|r|>0.5）
    sep_row = len(ALL_KEYS) + 4
    ws1.cell(row=sep_row, column=1, value='高相关策略对（|r| > 0.5）').font = BOLD
    ws1.cell(row=sep_row, column=1).fill = PatternFill('solid', start_color='FFF2CC')
    hdr(ws1, sep_row+1, 1, '策略A')
    hdr(ws1, sep_row+1, 2, '策略B')
    hdr(ws1, sep_row+1, 3, '相关系数')
    hdr(ws1, sep_row+1, 4, '公共月份数')
    hdr(ws1, sep_row+1, 5, '说明')

    pairs = []
    for i, ki in enumerate(ALL_KEYS):
        for j, kj in enumerate(ALL_KEYS):
            if j <= i: continue
            a, b, n = aligned_series(ki, kj)
            r = pearson(a, b)
            if not math.isnan(r) and abs(r) > 0.5:
                pairs.append((abs(r), ki, kj, r, n))
    pairs.sort(reverse=True)

    for idx, (_, ki, kj, r, n) in enumerate(pairs):
        r2 = sep_row + 2 + idx
        cell(ws1, r2, 1, ki, fill=fill_of(ki), align=LEFT)
        cell(ws1, r2, 2, kj, fill=fill_of(kj), align=LEFT)
        c = ws1.cell(row=r2, column=3, value=round(r, 3))
        c.number_format = '0.000'; c.alignment = CTR; c.border = BORD
        if r > 0.8:
            c.fill = PatternFill('solid', start_color='F4CCCC')
            c.font = Font(bold=True, name='Arial', size=10, color='9C0006')
        elif r > 0.6:
            c.fill = PatternFill('solid', start_color='FCE5CD')
            c.font = Font(name='Arial', size=10, color='B45F06')
        else:
            c.font = BODY
        cell(ws1, r2, 4, n)
        same_asset = asset_of(ki) == asset_of(kj) and asset_of(ki) != 'combo'
        note = '同标的同周期' if r > 0.9 else ('同标的' if same_asset else '跨标的')
        cell(ws1, r2, 5, note, align=LEFT)

    ws1.freeze_panes = 'C3'
    print('  Sheet1 相关性矩阵 完成')

    # ═══════════════════════════════════════════════════════════════════════════════
    # Sheet 2: 滚动12个月分析
    # ═══════════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('滚动12个月分析')

    ROLL_KEYS = ALL_KEYS
    ROLL_METRICS = ['夏普', '最大回撤%', '月胜率%']

    # 标题
    ws2.merge_cells('A1:D1')
    ws2['A1'] = '滚动12个月指标（夏普/最大回撤/月胜率）'
    ws2['A1'].font = Font(bold=True, name='Arial', size=12, color='1F4E79')
    ws2['A1'].alignment = CTR
    ws2['A1'].fill = PatternFill('solid', start_color='D6E4F0')

    # 汇总统计表
    hdr(ws2, 2, 1, '策略/组合', width=16)
    hdr(ws2, 2, 2, '类型', width=8)
    for ci, h in enumerate(['夏普均值','夏普最差','夏普最近','夏普<0占比',
                             '回撤均值%','回撤最大%','回撤最近%',
                             '胜率均值%','胜率最低%','胜率最近%',
                             '数据点数'], 3):
        hdr(ws2, 2, ci, h, width=11)

    for ri, k in enumerate(ROLL_KEYS, 3):
        rd = rolling_data.get(k)
        if not rd:
            continue
        sharpes  = [x[1] for x in rd]
        dds      = [x[2] for x in rd]
        wins     = [x[3] for x in rd]
        n        = len(rd)
        neg_pct  = sum(1 for s in sharpes if s < 0) / n * 100

        fill = fill_of(k)
        typ  = '单策略' if '_' in k else '融合'
        cell(ws2, ri, 1, k, fill=fill, align=LEFT, bold=True)
        cell(ws2, ri, 2, typ, fill=fill)

        vals = [
            (sum(sharpes)/n,    '0.00', sharpes, True),
            (min(sharpes),      '0.00', None,    True),
            (sharpes[-1],       '0.00', None,    True),
            (neg_pct,           '0.0"%"', None,  False),
            (sum(dds)/n,        '0.00"%"', dds,  False),
            (max(dds),          '0.00"%"', None, False),
            (dds[-1],           '0.00"%"', None, False),
            (sum(wins)/n,       '0.0"%"', wins,  True),
            (min(wins),         '0.0"%"', None,  True),
            (wins[-1],          '0.0"%"', None,  True),
            (n,                 '0',     None,   None),
        ]
        for ci, (v, fmt, _, _hb) in enumerate(vals, 3):
            c = ws2.cell(row=ri, column=ci, value=round(v, 3) if isinstance(v, float) else v)
            c.number_format = fmt; c.alignment = CTR; c.border = BORD; c.font = BODY
            if ri % 2 == 0: c.fill = PatternFill('solid', start_color='F5F5F5')

    # 条件格式：夏普均值列（C列）
    sharpe_range = f'E3:E{len(ROLL_KEYS)+2}'
    ws2.conditional_formatting.add(f'C3:C{len(ROLL_KEYS)+2}', ColorScaleRule(
        start_type='min', start_color='F4CCCC',
        mid_type='num',   mid_value=1.5, mid_color='FFFFBF',
        end_type='max',   end_color='D9EAD3',
    ))
    ws2.conditional_formatting.add(f'G3:G{len(ROLL_KEYS)+2}', ColorScaleRule(
        start_type='min', start_color='D9EAD3',
        end_type='max',   end_color='F4CCCC',
    ))

    # 时间序列明细（另起一块）
    sep2 = len(ROLL_KEYS) + 4
    ws2.cell(row=sep2, column=1, value='滚动指标时间序列明细').font = BOLD
    ws2.cell(row=sep2, column=1).fill = PatternFill('solid', start_color='FFF2CC')

    # 每个策略一组：日期 | 夏普 | 回撤 | 胜率
    col_offset = 1
    for k in ROLL_KEYS:
        rd = rolling_data.get(k)
        if not rd: continue
        # 组标题
        ws2.merge_cells(start_row=sep2+1, start_column=col_offset,
                        end_row=sep2+1, end_column=col_offset+3)
        c = ws2.cell(row=sep2+1, column=col_offset, value=k)
        c.fill = fill_of(k); c.font = BOLD; c.alignment = CTR; c.border = BORD
        for ci2, h in enumerate(['日期','夏普','回撤%','胜率%'], col_offset):
            hdr(ws2, sep2+2, ci2, h, width=9)
        for ri2, (dt, sh, dd, wr) in enumerate(rd, sep2+3):
            cell(ws2, ri2, col_offset,   dt.strftime('%Y-%m'), fmt='@')
            cell(ws2, ri2, col_offset+1, round(sh, 2), fmt='0.00')
            cell(ws2, ri2, col_offset+2, round(dd, 2), fmt='0.00"%"')
            cell(ws2, ri2, col_offset+3, round(wr, 1), fmt='0.0"%"')
        col_offset += 5

    ws2.freeze_panes = 'C3'
    print('  Sheet2 滚动分析 完成')

    # ═══════════════════════════════════════════════════════════════════════════════
    # Sheet 3: 月度收益明细表
    # ═══════════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet('月度收益明细')

    SHOW_KEYS = (list(files.keys()) +
                 [k for k in available_combos if k not in files])

    # 收集所有年月
    all_ym = set()
    for k in SHOW_KEYS:
        all_ym |= set(all_monthly.get(k, {}).keys())
    years  = sorted(set(y for y, m in all_ym))
    months = list(range(1, 13))

    # 标题
    total_cols = 1 + len(SHOW_KEYS)
    ws3.merge_cells(f'A1:{get_column_letter(total_cols)}1')
    ws3['A1'] = '月度净收益率 %（相对初始资金 $50,000）'
    ws3['A1'].font = Font(bold=True, name='Arial', size=12, color='1F4E79')
    ws3['A1'].alignment = CTR
    ws3['A1'].fill = PatternFill('solid', start_color='D6E4F0')

    # 列标题
    hdr(ws3, 2, 1, '年月', width=10)
    for ci, k in enumerate(SHOW_KEYS, 2):
        c = ws3.cell(row=2, column=ci, value=k)
        c.fill = fill_of(k); c.font = BOLD
        c.alignment = CTR; c.border = BORD
        ws3.column_dimensions[get_column_letter(ci)].width = 10

    # 数据行
    data_start = 3
    row_idx = data_start
    for y in years:
        for m in months:
            ym = (y, m)
            # 跳过没有任何数据的月份
            has_data = any(ym in all_monthly.get(k, {}) for k in SHOW_KEYS)
            if not has_data:
                continue
            label = f'{y}-{m:02d}'
            c = ws3.cell(row=row_idx, column=1, value=label)
            c.font = BOLD; c.alignment = CTR; c.border = BORD
            if y % 2 == 0:
                c.fill = PatternFill('solid', start_color='F5F5F5')

            for ci, k in enumerate(SHOW_KEYS, 2):
                pnl = all_monthly.get(k, {}).get(ym)
                c2 = ws3.cell(row=row_idx, column=ci)
                if pnl is not None:
                    pct = pnl / INITIAL * 100
                    c2.value = round(pct, 2)
                    c2.number_format = '0.00"%"'
                    # 颜色：正绿负红
                    if pct > 0:
                        intensity = min(int(pct / 5 * 100), 200)
                        g = hex(min(255, 180 + intensity // 3))[2:].zfill(2)
                        c2.fill = PatternFill('solid', start_color=f'C6EFC6')
                        c2.font = Font(name='Arial', size=10, color='375623')
                    elif pct < 0:
                        c2.fill = PatternFill('solid', start_color='FCE4D6')
                        c2.font = Font(name='Arial', size=10, color='9C0006')
                    else:
                        c2.font = BODY
                else:
                    c2.value = '-'
                    c2.font = Font(name='Arial', size=10, color='AAAAAA')
                c2.alignment = CTR; c2.border = BORD
            row_idx += 1

        # 年度小计行
        yr_row = row_idx
        c = ws3.cell(row=yr_row, column=1, value=f'{y} 合计')
        c.fill = PatternFill('solid', start_color='D9D9D9')
        c.font = Font(bold=True, name='Arial', size=10)
        c.alignment = CTR; c.border = BORD
        for ci, k in enumerate(SHOW_KEYS, 2):
            yr_pnl = sum(all_monthly.get(k, {}).get((y, m), 0.0) for m in months)
            yr_pct = yr_pnl / INITIAL * 100
            c2 = ws3.cell(row=yr_row, column=ci, value=round(yr_pct, 2))
            c2.number_format = '0.00"%"'
            c2.fill = PatternFill('solid', start_color='D9D9D9')
            c2.font = Font(bold=True, name='Arial', size=10,
                           color='375623' if yr_pct > 0 else ('9C0006' if yr_pct < 0 else '000000'))
            c2.alignment = CTR; c2.border = BORD
        row_idx += 1

    # 全期汇总行
    total_row = row_idx
    c = ws3.cell(row=total_row, column=1, value='全期合计')
    c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CTR; c.border = BORD
    for ci, k in enumerate(SHOW_KEYS, 2):
        total_pnl = sum(all_monthly.get(k, {}).values())
        total_pct = total_pnl / INITIAL * 100
        c2 = ws3.cell(row=total_row, column=ci, value=round(total_pct, 2))
        c2.number_format = '0.00"%"'
        c2.fill = HDR_FILL
        c2.font = Font(bold=True, name='Arial', size=10, color='FFFFFF')
        c2.alignment = CTR; c2.border = BORD

    # 月胜率汇总行
    wr_row = total_row + 1
    c = ws3.cell(row=wr_row, column=1, value='月胜率')
    c.fill = PatternFill('solid', start_color='FFF2CC')
    c.font = BOLD; c.alignment = CTR; c.border = BORD
    for ci, k in enumerate(SHOW_KEYS, 2):
        vals = list(all_monthly.get(k, {}).values())
        if vals:
            wr = sum(1 for v in vals if v > 0) / len(vals) * 100
            c2 = ws3.cell(row=wr_row, column=ci, value=round(wr, 1))
            c2.number_format = '0.0"%"'
            c2.fill = PatternFill('solid', start_color='FFF2CC')
            c2.font = BOLD; c2.alignment = CTR; c2.border = BORD

    ws3.freeze_panes = 'B3'
    ws3.row_dimensions[1].height = 20
    ws3.row_dimensions[2].height = 18
    print('  Sheet3 月度收益明细 完成')

    # ─── 保存 ─────────────────────────────────────────────────────────────────────
    # ─── 保存 ────────────────────────────────────────────────────────────
    import os as _os
    out_path = _os.path.join(out_dir, f'详细分析报告_{date_str}.xlsx')
    wb.save(out_path)
    print(f'\n已保存: {out_path}')
    print(f'包含 {len(wb.sheetnames)} 个 sheet: {wb.sheetnames}')


if __name__ == '__main__':
    import os, datetime
    base = os.path.expanduser('~/Downloads/')
    out  = os.path.expanduser('~/Downloads/')
    date_str = datetime.date.today().isoformat()
    run(DEFAULT_FILES, base='', out_dir=out, date_str=date_str)
