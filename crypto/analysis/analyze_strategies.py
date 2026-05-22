"""
策略综合分析脚本
生成指标对比 Excel 报告。

直接运行时使用 DEFAULT_FILES（需修改路径）；
由 analysis_core 调用时使用 run() 接口。
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as xl_numbers)
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
import datetime
import os

BASE = '/Users/huminghe/Downloads/'

FILES = {
    'BTC': {
        'ema':  'strategy_ema_btc_OKX_BTCUSDT.P_2026-05-18_bccdc.xlsx',
        'v2':   'v2_strategy_btc_OKX_BTCUSDT.P_2026-05-18_89f2b.xlsx',
        'v3':   'v3_205m_strategy_btc_OKX_BTCUSDT.P_2026-05-18_7f0bc.xlsx',
    },
    'ETH': {
        'ema':  'strategy_ema_eth_OKX_ETHUSDT.P_2026-05-18_6d950.xlsx',
        'v2':   'v2_strategy_eth_OKX_ETHUSDT.P_2026-05-18_7eba3.xlsx',
        'v3':   'v3_3h_strategy_eth_OKX_ETHUSDT.P_2026-05-18_f84bb.xlsx',
    },
    'SOL': {
        'ema':  'strategy_ema_sol_OKX_SOLUSDT.P_2026-05-18_f737f.xlsx',
        'v2':   'v2_strategy_sol_OKX_SOLUSDT.P_2026-05-18_2c3a0.xlsx',
        'v3':   'v3_3h_strategy_sol_OKX_SOLUSDT.P_2026-05-18_a51fb.xlsx',
    },
    'DOGE': {
        'ema':  'strategy_ema_meme_OKX_DOGEUSDT.P_2026-05-18_3a4e0.xlsx',
        'v2':   'v2_strategy_doge_OKX_DOGEUSDT.P_2026-05-18_2b9fa.xlsx',
        'v3':   'v3_205m_strategy_doge_OKX_DOGEUSDT.P_2026-05-18_283df.xlsx',
    },
}

ASSETS = ['BTC', 'ETH', 'SOL', 'DOGE']
STRATS = ['ema', 'v2', 'v3']
STRAT_LABELS = {'ema': 'EMA', 'v2': 'V2', 'v3': 'V3'}

# ─── 样式 ────────────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill('solid', start_color='1F4E79')
HDR_FONT   = Font(bold=True, color='FFFFFF', name='Arial', size=10)
SUB_FILL   = PatternFill('solid', start_color='2E75B6')
SUB_FONT   = Font(bold=True, color='FFFFFF', name='Arial', size=10)
BEST_FILL  = PatternFill('solid', start_color='E2EFDA')
WARN_FILL  = PatternFill('solid', start_color='FCE4D6')
ALT_FILL   = PatternFill('solid', start_color='F2F2F2')
BODY_FONT  = Font(name='Arial', size=10)
BOLD_FONT  = Font(bold=True, name='Arial', size=10)
CENTER     = Alignment(horizontal='center', vertical='center')
LEFT       = Alignment(horizontal='left',   vertical='center')
RIGHT      = Alignment(horizontal='right',  vertical='center')

THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ASSET_COLORS = {
    'BTC':  'F7931A',
    'ETH':  '627EEA',
    'SOL':  '9945FF',
    'DOGE': 'C2A633',
}

def cell_style(ws, cell_ref, fill=None, font=None, align=None, border=None, num_fmt=None):
    c = ws[cell_ref] if isinstance(cell_ref, str) else cell_ref
    if fill:   c.fill      = fill
    if font:   c.font      = font
    if align:  c.alignment = align
    if border: c.border    = border
    if num_fmt: c.number_format = num_fmt

def set_row_style(ws, row, cols, fill=None, font=None, align=None, border=None):
    for col in range(1, cols + 1):
        c = ws.cell(row=row, column=col)
        if fill:   c.fill      = fill
        if font:   c.font      = font
        if align:  c.alignment = align
        if border: c.border    = border

# ─── 数据读取 ─────────────────────────────────────────────────────────────────
def read_stats(filepath):
    wb = openpyxl.load_workbook(filepath, read_only=True)
    stats = {}
    for sheet_name in ['表现', '交易分析', '风险调整后的表现']:
        ws = wb[sheet_name]
        for r in ws.iter_rows(values_only=True):
            if r[0] is not None:
                stats[r[0]] = {
                    '全部USDT': r[1], '全部%': r[2],
                    '多头USDT': r[3], '多头%': r[4],
                    '空头USDT': r[5], '空头%': r[6],
                }
    wb.close()
    return stats

def read_equity_curve(filepath):
    """返回 [(datetime, cumulative_pnl_pct), ...] 出场时间序列"""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb['交易清单']
    points = []
    for r in ws.iter_rows(values_only=True):
        if r[1] and '出场' in str(r[1]) and isinstance(r[2], datetime.datetime) and r[14] is not None:
            points.append((r[2], float(r[14])))
    wb.close()
    return sorted(points, key=lambda x: x[0])

def get_key(stats, key, field='全部%', default=None):
    return stats.get(key, {}).get(field, default)

if __name__ == "__main__":
    import os
    _base = os.path.expanduser("~/Downloads/")
    _files = {f"{a}_{s}": _base + FILES[a][s] for a in ASSETS for s in STRATS}
    run(_files, base="", out_dir=_base)


# ─── 供 run_analysis.py 调用的接口 ────────────────────────────────────────────

METRICS = [
    ('净利润 %',        '净利润',           '全部%',    '%.2f%%',  True),
    ('年化收益 %',      '年化收益率（CAGR）','全部%',    '%.2f%%',  True),
    ('最大回撤 %',      '最大股权回撤（intrabar）','全部%','%.2f%%', False),
    ('夏普比率',        '夏普比率',          '全部USDT', '%.3f',    True),
    ('Sortino比率',     'Sortino比率',       '全部USDT', '%.3f',    True),
    ('盈利因子',        '盈利因子',          '全部USDT', '%.3f',    True),
    ('胜率 %',          '获利百分比',        '全部%',    '%.2f%%',  True),
    ('盈亏比',          '平均胜率/平均负率', '全部USDT', '%.3f',    True),
    ('总交易次数',      '总交易',            '全部USDT', '%d',      None),
    ('多头净利润 %',    '净利润',            '多头%',    '%.2f%%',  True),
    ('空头净利润 %',    '净利润',            '空头%',    '%.2f%%',  True),
    ('多头盈利因子',    '盈利因子',          '多头USDT', '%.3f',    True),
    ('空头盈利因子',    '盈利因子',          '空头USDT', '%.3f',    True),
]


def run(files: dict, base: str, out_dir: str, date_str: str = None):
    """
    files: {key: 绝对路径}，key 格式 ASSET_strat
    base:  路径前缀（run_analysis 传空字符串）
    """
    import datetime as _dt, os
    if date_str is None:
        date_str = _dt.date.today().isoformat()

    # 重建 ASSETS / strats_per_asset
    _assets = sorted(set(k.split("_")[0] for k in files))
    _strats_per: dict[str, list[str]] = {}
    for k in files:
        a, s = k.split("_", 1)
        _strats_per.setdefault(a, []).append(s)

    _all_stats: dict = {}
    _all_equity: dict = {}
    for k, fp in files.items():
        a, s = k.split("_", 1)
        _all_stats.setdefault(a, {})[s] = read_stats(base + fp)
        _all_equity.setdefault(a, {})[s] = read_equity_curve(base + fp)

    # 统一时间轴
    _all_dates_set = set()
    for a in _assets:
        for s in _strats_per.get(a, []):
            for dt, _ in _all_equity[a][s]:
                _all_dates_set.add(dt)
    _all_dates = sorted(_all_dates_set)

    # 插值
    def _interp(pts, dates):
        j = 0
        result = []
        for d in dates:
            while j < len(pts) - 1 and pts[j+1][0] <= d:
                j += 1
            result.append(pts[j][1] if pts[j][0] <= d else 0.0)
        return result

    _series = {}
    for k, fp in files.items():
        a, s = k.split("_", 1)
        pts = [(p[0], p[1]) for p in _all_equity[a][s]]
        _series[k] = _interp(pts, _all_dates)

    # 各标的融合
    _merged = {}
    for a in _assets:
        strats = _strats_per.get(a, [])
        keys = [f"{a}_{s}" for s in strats]
        _merged[a] = [sum(_series[k][i] for k in keys) / len(keys)
                      for i in range(len(_all_dates))]

    # 全组合
    _portfolio = [sum(_merged[a][i] for a in _assets) / len(_assets)
                  for i in range(len(_all_dates))]

    # 月度采样
    _monthly_dates = []
    _prev_ym = None
    for d in _all_dates:
        ym = (d.year, d.month)
        if ym != _prev_ym:
            _monthly_dates.append(d)
            _prev_ym = ym

    def _get_val_at(dm, d):
        best = None
        for dd, v in sorted(dm.items()):
            if dd <= d:
                best = v
        return best or 0.0

    # 构建 Excel
    from openpyxl import Workbook as _WB
    from openpyxl.utils import get_column_letter as _gcl
    import datetime as _dt2

    _wb = _WB()
    _wb.remove(_wb.active)

    # Sheet1: 总览对比
    _ws1 = _wb.create_sheet("总览对比")
    _headers = ["指标"] + [f"{a}_{s}" for a in _assets for s in _strats_per.get(a, [])]
    for ci, h in enumerate(_headers, 1):
        c = _ws1.cell(row=1, column=ci, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = CENTER; c.border = BORDER
    for mi, (label, key, field, fmt, hb) in enumerate(METRICS):
        row = mi + 2
        _ws1.cell(row=row, column=1, value=label).font = BOLD_FONT
        ci = 2
        vals = {}
        for a in _assets:
            for s in _strats_per.get(a, []):
                v = get_key(_all_stats.get(a, {}).get(s, {}), key, field)
                if v is not None:
                    vals[(a, s)] = float(v)
        best_v = (max(vals.values()) if hb else min(vals.values())) if vals and hb is not None else None
        for a in _assets:
            for s in _strats_per.get(a, []):
                v = vals.get((a, s))
                c = _ws1.cell(row=row, column=ci)
                if v is not None:
                    c.value = v
                    c.number_format = '0.00"%"' if '%' in fmt else '0.000'
                    if best_v is not None and abs(v - best_v) < 1e-9:
                        c.fill = BEST_FILL
                        c.font = Font(bold=True, name='Arial', size=10, color='375623')
                    else:
                        c.font = BODY_FONT
                else:
                    c.value = '-'; c.font = BODY_FONT
                c.alignment = CENTER; c.border = BORDER
                ci += 1
    _ws1.column_dimensions['A'].width = 22
    for ci in range(2, len(_headers) + 1):
        _ws1.column_dimensions[_gcl(ci)].width = 11
    _ws1.freeze_panes = 'B2'

    # Sheet2: 月度收益明细
    _ws2 = _wb.create_sheet("月度收益明细")
    _h2 = ["年月"] + [f"{a}融合%" for a in _assets] + ["全组合%"]
    for ci, h in enumerate(_h2, 1):
        c = _ws2.cell(row=1, column=ci, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = CENTER; c.border = BORDER
    _merged_maps = {a: {d: v for d, v in zip(_all_dates, _merged[a])} for a in _assets}
    _port_map = {d: v for d, v in zip(_all_dates, _portfolio)}
    for ri, d in enumerate(_monthly_dates, 2):
        _ws2.cell(row=ri, column=1, value=f"{d.year}-{d.month:02d}")
        for ci, a in enumerate(_assets, 2):
            v = _get_val_at(_merged_maps[a], d)
            c = _ws2.cell(row=ri, column=ci, value=round(v, 2))
            c.number_format = '0.00"%"'; c.alignment = CENTER; c.border = BORDER
        v = _get_val_at(_port_map, d)
        c = _ws2.cell(row=ri, column=len(_assets) + 2, value=round(v, 2))
        c.number_format = '0.00"%"'; c.alignment = CENTER; c.border = BORDER
    _ws2.freeze_panes = 'A2'

    _out = os.path.join(out_dir, f"策略综合分析报告_{date_str}.xlsx")
    _wb.save(_out)
    print(f"  Excel 报告已保存: {_out}")

