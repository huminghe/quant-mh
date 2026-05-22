"""
生成收益走势交互图（HTML）
"""
import openpyxl
import datetime
import plotly.graph_objects as go
from plotly.subplots import make_subplots

ASSET_COLORS = {
    "BTC":  "#F7931A",
    "ETH":  "#627EEA",
    "SOL":  "#9945FF",
    "DOGE": "#C2A633",
}
STRAT_DASH   = {"ema": "dot", "v2": "dash", "v3": "solid"}
STRAT_LABELS = {"ema": "EMA", "v2": "V2",  "v3": "V3"}


def read_equity_curve(filepath):
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb["交易清单"]
    points = []
    for r in ws.iter_rows(values_only=True):
        if r[1] and "出场" in str(r[1]) and isinstance(r[2], datetime.datetime) and r[14] is not None:
            points.append((r[2], float(r[14])))
    wb.close()
    return sorted(points, key=lambda x: x[0])


def read_stats(filepath):
    wb = openpyxl.load_workbook(filepath, read_only=True)
    stats = {}
    for sheet_name in ["表现", "交易分析", "风险调整后的表现"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            if row[0] is not None:
                stats[row[0]] = {
                    "全部USDT": row[1], "全部%": row[2],
                    "多头USDT": row[3], "多头%": row[4],
                    "空头USDT": row[5], "空头%": row[6],
                }
    wb.close()
    return stats


def gk(stats, key, field="全部%", default=0):
    return float(stats.get(key, {}).get(field) or default)


if __name__ == "__main__":
    pass  # 直接运行时请调用 run()


# ─── 供 run_analysis.py 调用的接口 ────────────────────────────────────────────

def run(files: dict, base: str, out_dir: str):
    """files: {key: 绝对路径}，base 传空字符串"""
    import os

    _assets = sorted(set(k.split("_")[0] for k in files))
    _strats_per: dict[str, list[str]] = {}
    for k in files:
        a, s = k.split("_", 1)
        _strats_per.setdefault(a, []).append(s)

    # 读取曲线
    _all_equity: dict = {}
    for k, fp in files.items():
        a, s = k.split("_", 1)
        _all_equity.setdefault(a, {})[s] = read_equity_curve(base + fp)

    # 统一时间轴
    _all_dates_set = set()
    for a in _assets:
        for s in _strats_per.get(a, []):
            for dt, _ in _all_equity[a][s]:
                _all_dates_set.add(dt.date() if hasattr(dt, 'date') else dt)
    _all_dates = sorted(_all_dates_set)

    def _interp(pts, dates):
        dm = {}
        for dt, v in pts:
            d = dt.date() if hasattr(dt, 'date') else dt
            dm[d] = v
        sorted_pts = sorted(dm.items())
        j = 0
        result = []
        for d in dates:
            while j < len(sorted_pts) - 1 and sorted_pts[j+1][0] <= d:
                j += 1
            result.append(sorted_pts[j][1] if sorted_pts[j][0] <= d else 0.0)
        return result

    _series = {}
    for k, fp in files.items():
        a, s = k.split("_", 1)
        _series[k] = _interp(_all_equity[a][s], _all_dates)

    # 各标的融合
    _merged = {}
    for a in _assets:
        keys = [f"{a}_{s}" for s in _strats_per.get(a, [])]
        _merged[a] = [sum(_series[k][i] for k in keys) / len(keys)
                      for i in range(len(_all_dates))]

    _portfolio = [sum(_merged[a][i] for a in _assets) / len(_assets)
                  for i in range(len(_all_dates))]

    # 月度采样
    _monthly_dates = []
    _prev_ym = None
    for d in _all_dates:
        ym = (d.year, d.month)
        if ym != _prev_ym:
            _monthly_dates.append(d)
            _prev_ym = ym

    def _get_val_at(dm, d):
        best = None
        for dd, v in sorted(dm.items()):
            if dd <= d:
                best = v
        return best or 0.0

    # 图1：各标的三策略对比
    _fig1 = make_subplots(
        rows=2, cols=2,
        subplot_titles=[f'{a} — 三策略收益曲线' for a in _assets],
        shared_xaxes=False,
        vertical_spacing=0.12,
        horizontal_spacing=0.08,
    )
    _pos = {a: ((i // 2) + 1, (i % 2) + 1) for i, a in enumerate(_assets)}
    for a in _assets:
        r, c = _pos[a]
        color = ASSET_COLORS.get(a, '#888')
        for s in _strats_per.get(a, []):
            pts = _all_equity[a][s]
            dates = [p[0] for p in pts]
            vals  = [p[1] for p in pts]
            _fig1.add_trace(go.Scatter(
                x=dates, y=vals, name=f'{a} {s.upper()}',
                line=dict(color=color, dash=STRAT_DASH.get(s, 'solid'), width=1.5),
                opacity=0.85, legendgroup=f'{a}_{s}',
            ), row=r, col=c)
        m_dates = _all_dates
        m_vals  = _merged[a]
        _fig1.add_trace(go.Scatter(
            x=m_dates, y=m_vals, name=f'{a} 融合',
            line=dict(color=color, width=3),
            legendgroup=f'{a}_融合',
        ), row=r, col=c)
    _fig1.update_layout(
        title=dict(text='各标的三策略收益走势对比（累计P&L %）', font=dict(size=16)),
        height=900, template='plotly_white', hovermode='x unified',
    )
    _fig1.write_html(os.path.join(out_dir, '图1_各标的策略对比.html'))
    print("  图1 完成")

    # 图2：融合曲线 + 全组合
    _fig2 = go.Figure()
    for a in _assets:
        _fig2.add_trace(go.Scatter(
            x=_all_dates, y=_merged[a], name=f'{a} 融合',
            line=dict(color=ASSET_COLORS.get(a, '#888'), width=2),
        ))
    _fig2.add_trace(go.Scatter(
        x=_all_dates, y=_portfolio, name='全组合（等权）',
        line=dict(color='#2E75B6', width=3),
        fill='tozeroy', fillcolor='rgba(46,117,182,0.08)',
    ))
    _fig2.update_layout(
        title=dict(text='各标的融合策略 & 全组合收益走势（累计P&L %）', font=dict(size=16)),
        height=550, template='plotly_white', hovermode='x unified',
        yaxis=dict(ticksuffix='%'),
        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
    )
    _fig2.write_html(os.path.join(out_dir, '图2_融合策略与全组合.html'))
    print("  图2 完成")

    # 图3：关键指标柱状对比（需要 stats）
    _all_stats: dict = {}
    for k, fp in files.items():
        a, s = k.split("_", 1)
        _all_stats.setdefault(a, {})[s] = read_stats(base + fp)

    _fig3 = make_subplots(
        rows=2, cols=2,
        subplot_titles=['净利润 %', '夏普比率', '最大回撤 %', '盈利因子'],
        vertical_spacing=0.18, horizontal_spacing=0.1,
    )
    _metric_configs = [
        ('净利润 %',   '净利润',                    '全部%',    (1, 1)),
        ('夏普比率',   '夏普比率',                   '全部USDT', (1, 2)),
        ('最大回撤 %', '最大股权回撤（intrabar）',   '全部%',    (2, 1)),
        ('盈利因子',   '盈利因子',                   '全部USDT', (2, 2)),
    ]
    _strat_colors = {'ema': '#4472C4', 'v2': '#ED7D31', 'v3': '#70AD47'}
    _all_strats = sorted(set(k.split("_", 1)[1] for k in files))
    for label, key, field, (r, c) in _metric_configs:
        for s in _all_strats:
            y_vals = [gk(_all_stats.get(a, {}).get(s, {}), key, field)
                      for a in _assets]
            _fig3.add_trace(go.Bar(
                name=s.upper(), x=_assets, y=y_vals,
                marker_color=_strat_colors.get(s, '#888'),
                showlegend=(r == 1 and c == 1),
                legendgroup=s,
            ), row=r, col=c)
    _fig3.update_layout(
        title=dict(text='各标的 × 各策略关键指标对比', font=dict(size=16)),
        height=700, template='plotly_white', barmode='group',
        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
    )
    _fig3.layout['yaxis'].ticksuffix = '%'
    _fig3.layout['yaxis3'].ticksuffix = '%'
    _fig3.write_html(os.path.join(out_dir, '图3_关键指标柱状对比.html'))
    print("  图3 完成")
