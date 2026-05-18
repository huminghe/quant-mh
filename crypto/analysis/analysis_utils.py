"""
分析工具公共模块
提供数据读取、插值、回撤计算等共用函数
"""
import datetime
import openpyxl


def read_equity_curve(filepath: str) -> list[tuple[datetime.date, float]]:
    """
    从 TradingView 导出的 Excel 读取出场时间序列。
    返回 [(date, cumulative_pnl_pct), ...] 按日期升序。
    """
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb["交易清单"]
    dm: dict[datetime.date, float] = {}
    for r in ws.iter_rows(values_only=True):
        if (r[1] and "出场" in str(r[1])
                and isinstance(r[2], datetime.datetime)
                and r[14] is not None):
            dm[r[2].date()] = float(r[14])
    wb.close()
    return sorted(dm.items())


def read_monthly_pnl(filepath: str) -> dict[tuple[int, int], float]:
    """
    从 Excel 交易清单提取月度净损益 USDT。
    返回 {(year, month): pnl_usdt}
    """
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb["交易清单"]
    monthly: dict[tuple[int, int], float] = {}
    for r in ws.iter_rows(values_only=True):
        if (r[1] and "出场" in str(r[1])
                and isinstance(r[2], datetime.datetime)
                and r[7] is not None):
            ym = (r[2].year, r[2].month)
            monthly[ym] = monthly.get(ym, 0.0) + float(r[7])
    wb.close()
    return monthly


def interp_series(pts: list[tuple[datetime.date, float]],
                  dates: list[datetime.date]) -> list[float]:
    """将稀疏的 (date, value) 序列前向填充到 dates 列表。"""
    result = []
    j = 0
    for d in dates:
        while j < len(pts) - 1 and pts[j + 1][0] <= d:
            j += 1
        result.append(pts[j][1] if pts[j][0] <= d else 0.0)
    return result


def max_drawdown(vals: list[float]) -> float:
    """计算最大回撤（百分点）。"""
    peak = vals[0]
    md = 0.0
    for v in vals:
        if v > peak:
            peak = v
        md = max(md, peak - v)
    return md


def drawdown_series(vals: list[float]) -> list[float]:
    """返回每个时间点相对历史峰值的回撤（负值）。"""
    peak = vals[0]
    result = []
    for v in vals:
        if v > peak:
            peak = v
        result.append(-(peak - v))
    return result


def avg_series(*series: list[float]) -> list[float]:
    """多条等长序列等权平均。"""
    n = len(series)
    return [sum(s[i] for s in series) / n for i in range(len(series[0]))]
