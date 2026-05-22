"""
生成 Markdown 结论文档，内嵌关键图表的 PNG 截图。

依赖 kaleido（pip install kaleido）用于 plotly figure → PNG。
"""
import datetime
import math
import os
from pathlib import Path

from analysis_utils import read_equity_curve, read_monthly_pnl, interp_series
import plotly.graph_objects as go
from plotly.subplots import make_subplots


INITIAL = 50_000.0

# 默认颜色（不在此表中的 key 自动分配颜色）
ASSET_COLORS = {
    "BTC": "#F7931A", "ETH": "#627EEA",
    "SOL": "#9945FF", "DOGE": "#C2A633",
}

# 自动颜色池（用于版本分析模式）
AUTO_COLORS = [
    "#4472C4", "#ED7D31", "#70AD47", "#FFC000",
    "#5B9BD5", "#FF0000", "#00B0F0", "#7030A0",
]


def _asset_color(key: str, idx: int = 0) -> str:
    """根据 key 前缀匹配颜色，找不到则用自动颜色池。"""
    prefix = key.split("_")[0].upper()
    return ASSET_COLORS.get(prefix, AUTO_COLORS[idx % len(AUTO_COLORS)])


def _drawdown_series(vals: list[float]) -> list[float]:
    peak = vals[0] if vals else 0.0
    result = []
    for v in vals:
        if v > peak:
            peak = v
        result.append(-(peak - v))
    return result


def _max_dd(vals: list[float]) -> float:
    peak = vals[0] if vals else 0.0
    md = 0.0
    for v in vals:
        if v > peak:
            peak = v
        md = max(md, peak - v)
    return md


def _sharpe(monthly_pnl: dict) -> float:
    if not monthly_pnl:
        return 0.0
    vals = [v / INITIAL for v in monthly_pnl.values()]
    mean = sum(vals) / len(vals)
    std = math.sqrt(sum((v - mean) ** 2 for v in vals) / len(vals))
    return mean / std * math.sqrt(12) if std > 1e-9 else 0.0


def _sortino(monthly_pnl: dict) -> float:
    if not monthly_pnl:
        return 0.0
    vals = [v / INITIAL for v in monthly_pnl.values()]
    mean = sum(vals) / len(vals)
    neg = [v for v in vals if v < 0]
    downstd = math.sqrt(sum(v ** 2 for v in neg) / len(neg)) if neg else 1e-9
    return mean / downstd * math.sqrt(12) if downstd > 1e-9 else 0.0


def _pearson(a: list[float], b: list[float]) -> float:
    n = len(a)
    if n < 2:
        return float("nan")
    ma, mb = sum(a) / n, sum(b) / n
    cov = sum((a[i] - ma) * (b[i] - mb) for i in range(n))
    sa = math.sqrt(sum((x - ma) ** 2 for x in a))
    sb = math.sqrt(sum((x - mb) ** 2 for x in b))
    return cov / (sa * sb) if sa > 1e-9 and sb > 1e-9 else float("nan")


def _save_png(fig: go.Figure, path: str, width: int = 1200, height: int = 500):
    try:
        fig.write_image(path, width=width, height=height, scale=1.5)
        print(f"  PNG: {Path(path).name}")
    except Exception as e:
        print(f"  PNG 导出失败 ({Path(path).name}): {e}")


def _read_stats_simple(fp: str) -> dict:
    """读取 xlsx 表现/风险调整后的表现 sheet，返回 {指标名: {全部USDT, 全部%}}。
    兼容中文（新版）和英文（旧版）sheet 名。
    """
    import openpyxl
    wb = openpyxl.load_workbook(fp, read_only=True)
    stats = {}
    # 中文名 → 英文名 fallback
    sheet_map = [
        ("表现",              "Performance"),
        ("交易分析",          "Trades analysis"),
        ("风险调整后的表现",  "Risk performance ratios"),
    ]
    for cn, en in sheet_map:
        sn = cn if cn in wb.sheetnames else (en if en in wb.sheetnames else None)
        if sn is None:
            continue
        ws = wb[sn]
        for r in ws.iter_rows(values_only=True):
            if r[0] is not None:
                stats[r[0]] = {"全部USDT": r[1], "全部%": r[2]}
    wb.close()
    return stats


def _read_trades(fp: str) -> list[dict]:
    """读取交易清单，返回已平仓交易列表（每笔含 year, profit_usdt, bars）。
    兼容中文（交易清单）和英文（List of trades）sheet 名。
    """
    import openpyxl
    import datetime as _dt
    wb = openpyxl.load_workbook(fp, read_only=True)
    sn = ("交易清单" if "交易清单" in wb.sheetnames
          else ("List of trades" if "List of trades" in wb.sheetnames else None))
    if sn is None:
        wb.close()
        return []

    ws = wb[sn]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    # 找表头行（第一行含 Trade # 或 交易#）
    header = rows[0]
    col_map = {}
    for i, h in enumerate(header):
        if h is None:
            continue
        hs = str(h).strip()
        if hs in ("Trade #", "交易#", "交易 #"):
            col_map["trade_no"] = i
        elif hs in ("Type", "类型"):
            col_map["type"] = i
        elif hs in ("Date/Time", "日期/时间", "日期和时间"):
            col_map["dt"] = i
        elif hs in ("Profit USDT", "盈亏 USDT", "盈亏USDT", "净损益 USDT", "净损益USDT"):
            col_map["profit"] = i
        elif hs in ("Avg # bars in trades", "交易的平均#K线数"):
            col_map["bars"] = i

    trades = []
    for r in rows[1:]:
        if not r or r[0] is None:
            continue
        # 只取 Exit 行（一笔交易的结果行）
        t = str(r[col_map.get("type", 1)] or "").lower()
        if "exit" not in t and "平仓" not in t and "出场" not in t:
            continue
        dt_val = r[col_map.get("dt", 3)]
        if isinstance(dt_val, _dt.datetime):
            year = dt_val.year
        else:
            continue
        profit = float(r[col_map.get("profit", 6)] or 0.0)
        trades.append({"year": year, "profit": profit})

    return trades


def _annual_pnl(trades: list[dict]) -> dict[int, float]:
    """按年汇总净利润（USDT）。"""
    result: dict[int, float] = {}
    for t in trades:
        result[t["year"]] = result.get(t["year"], 0.0) + t["profit"]
    return result


def _max_consec_loss(trades: list[dict]) -> int:
    """最大连续亏损笔数。"""
    max_streak = cur = 0
    for t in trades:
        if t["profit"] < 0:
            cur += 1
            max_streak = max(max_streak, cur)
        else:
            cur = 0
    return max_streak


def _gk(stats: dict, key: str, field: str = "全部%", default: float = 0.0) -> float:
    """从 stats dict 安全取值。"""
    return float(stats.get(key, {}).get(field) or default)


# ─── 核心计算 ─────────────────────────────────────────────────────────────────

def _compute(files: dict[str, str]) -> dict:
    """读取所有文件，计算曲线、融合、组合等数据。"""
    assets_set = sorted(set(k.split("_")[0] for k in files))
    strats_per: dict[str, list[str]] = {}
    for k in files:
        a, s = k.split("_", 1)
        strats_per.setdefault(a, []).append(s)

    single_asset_mode = len(assets_set) == 1

    # 读取权益曲线、月度 P&L、stats、交易清单
    equity: dict[str, list] = {}
    monthly: dict[str, dict] = {}
    raw_stats: dict[str, dict] = {}
    trades: dict[str, list] = {}
    for k, fp in files.items():
        equity[k] = read_equity_curve(fp)
        monthly[k] = read_monthly_pnl(fp)
        raw_stats[k] = _read_stats_simple(fp)
        trades[k] = _read_trades(fp)

    # 统一时间轴
    all_dates = sorted(set(d for pts in equity.values() for d, _ in pts))

    # 插值序列
    series: dict[str, list[float]] = {}
    for k, pts in equity.items():
        series[k] = interp_series(pts, all_dates)

    def _avg_series(*arrs):
        n = len(arrs)
        return [sum(a[i] for a in arrs) / n for i in range(len(arrs[0]))]

    def _avg_monthly(*mdicts):
        """等权合并多个月度 P&L dict。"""
        n = len(mdicts)
        all_ym: set = set()
        for m in mdicts:
            all_ym |= set(m.keys())
        return {ym: sum(m.get(ym, 0.0) for m in mdicts) / n for ym in all_ym}

    combos: dict[str, list[float]] = {}
    combo_monthly: dict[str, dict] = {}   # 所有组合（含融合）的月度 P&L
    merged: dict[str, list[float]] = {}
    merged_monthly: dict[str, dict] = {}

    if single_asset_mode:
        # 单标的：每个版本独立，直接做版本间组合
        a = assets_set[0]
        strats = strats_per.get(a, [])
        units = {s: series[f"{a}_{s}"] for s in strats if f"{a}_{s}" in series}
        units_m = {s: monthly[f"{a}_{s}"] for s in strats if f"{a}_{s}" in monthly}
        unit_list = list(units.keys())
        combos.update(units)
        for s in unit_list:
            combo_monthly[s] = units_m.get(s, {})
        for i in range(len(unit_list)):
            for j in range(i + 1, len(unit_list)):
                u1, u2 = unit_list[i], unit_list[j]
                name = f"{u1}+{u2}"
                combos[name] = _avg_series(units[u1], units[u2])
                combo_monthly[name] = _avg_monthly(units_m[u1], units_m[u2])
        for i in range(len(unit_list)):
            for j in range(i + 1, len(unit_list)):
                for kk in range(j + 1, len(unit_list)):
                    u1, u2, u3 = unit_list[i], unit_list[j], unit_list[kk]
                    name = f"{u1}+{u2}+{u3}"
                    combos[name] = _avg_series(units[u1], units[u2], units[u3])
                    combo_monthly[name] = _avg_monthly(units_m[u1], units_m[u2], units_m[u3])
        if len(unit_list) >= 4:
            combos["全版本等权"] = _avg_series(*list(units.values()))
            combo_monthly["全版本等权"] = _avg_monthly(*list(units_m.values()))
    else:
        # 多标的：先按标的融合，再跨标的组合
        for a in assets_set:
            keys = [f"{a}_{s}" for s in strats_per.get(a, []) if f"{a}_{s}" in series]
            if keys:
                merged[a] = _avg_series(*[series[k] for k in keys])
                combos[f"{a}融合"] = merged[a]
        asset_list = list(merged.keys())
        for i in range(len(asset_list)):
            for j in range(i + 1, len(asset_list)):
                a1, a2 = asset_list[i], asset_list[j]
                name = f"{a1}+{a2}"
                combos[name] = _avg_series(merged[a1], merged[a2])
        for i in range(len(asset_list)):
            for j in range(i + 1, len(asset_list)):
                for k in range(j + 1, len(asset_list)):
                    a1, a2, a3 = asset_list[i], asset_list[j], asset_list[k]
                    name = f"{a1}+{a2}+{a3}"
                    combos[name] = _avg_series(merged[a1], merged[a2], merged[a3])
        if len(asset_list) >= 4:
            combos["全4标的"] = _avg_series(*[merged[a] for a in asset_list])

        # 月度融合 P&L（标的级）
        for a in assets_set:
            keys = [f"{a}_{s}" for s in strats_per.get(a, []) if f"{a}_{s}" in monthly]
            if keys:
                all_ym: set = set()
                for k in keys:
                    all_ym |= set(monthly[k].keys())
                merged_monthly[a] = {
                    ym: sum(monthly[k].get(ym, 0.0) for k in keys) / len(keys)
                    for ym in all_ym
                }

        # 组合级月度 P&L
        for a in assets_set:
            fk = f"{a}融合"
            if a in merged_monthly:
                combo_monthly[fk] = merged_monthly[a]
        for i in range(len(asset_list)):
            for j in range(i + 1, len(asset_list)):
                a1, a2 = asset_list[i], asset_list[j]
                name = f"{a1}+{a2}"
                if a1 in merged_monthly and a2 in merged_monthly:
                    combo_monthly[name] = _avg_monthly(merged_monthly[a1], merged_monthly[a2])
        for i in range(len(asset_list)):
            for j in range(i + 1, len(asset_list)):
                for k in range(j + 1, len(asset_list)):
                    a1, a2, a3 = asset_list[i], asset_list[j], asset_list[k]
                    name = f"{a1}+{a2}+{a3}"
                    if all(x in merged_monthly for x in [a1, a2, a3]):
                        combo_monthly[name] = _avg_monthly(
                            merged_monthly[a1], merged_monthly[a2], merged_monthly[a3])
        if len(asset_list) >= 4:
            if all(a in merged_monthly for a in asset_list):
                combo_monthly["全4标的"] = _avg_monthly(*[merged_monthly[a] for a in asset_list])

    return {
        "assets": assets_set,
        "strats_per": strats_per,
        "single_asset_mode": single_asset_mode,
        "all_dates": all_dates,
        "series": series,
        "merged": merged,
        "combos": combos,
        "combo_monthly": combo_monthly,
        "monthly": monthly,
        "merged_monthly": merged_monthly,
        "raw_stats": raw_stats,
        "trades": trades,
    }


# ─── 图表生成 ─────────────────────────────────────────────────────────────────

def _fig_equity_overview(data: dict) -> go.Figure:
    """图1：各策略收益曲线总览。"""
    series = data["series"]
    merged = data["merged"]
    all_dates = data["all_dates"]
    assets = data["assets"]
    strats_per = data["strats_per"]

    n = len(assets)
    cols = min(n, 2)
    rows = math.ceil(n / cols) if n > 1 else 1

    if n == 1:
        fig = go.Figure()
        a = assets[0]
        color = _asset_color(a)
        for idx, s in enumerate(strats_per.get(a, [])):
            k = f"{a}_{s}"
            fig.add_trace(go.Scatter(
                x=all_dates, y=series[k],
                name=s.upper(),
                line=dict(color=AUTO_COLORS[idx % len(AUTO_COLORS)], width=1.8),
            ))
        if a in merged:
            fig.add_trace(go.Scatter(
                x=all_dates, y=merged[a],
                name="融合", line=dict(color=color, width=3),
            ))
        fig.update_layout(
            title=f"{a} — 各版本收益曲线",
            height=450, template="plotly_white", hovermode="x unified",
            yaxis=dict(ticksuffix="%"),
        )
    else:
        fig = make_subplots(
            rows=rows, cols=cols,
            subplot_titles=[f"{a} — 各策略收益" for a in assets],
            vertical_spacing=0.12, horizontal_spacing=0.08,
        )
        for i, a in enumerate(assets):
            r, c = i // cols + 1, i % cols + 1
            color = _asset_color(a, i)
            for idx, s in enumerate(strats_per.get(a, [])):
                k = f"{a}_{s}"
                fig.add_trace(go.Scatter(
                    x=all_dates, y=series[k],
                    name=f"{a} {s.upper()}",
                    line=dict(color=color,
                              dash=["dot", "dash", "solid", "dashdot"][idx % 4],
                              width=1.5),
                    legendgroup=a, opacity=0.8,
                ), row=r, col=c)
            if a in merged:
                fig.add_trace(go.Scatter(
                    x=all_dates, y=merged[a],
                    name=f"{a} 融合",
                    line=dict(color=color, width=3),
                    legendgroup=a,
                ), row=r, col=c)
        fig.update_layout(
            title="各标的策略收益曲线对比（累计 P&L %）",
            height=max(450, rows * 350),
            template="plotly_white", hovermode="x unified",
        )
        for ax in fig.layout:
            if ax.startswith("yaxis"):
                fig.layout[ax].ticksuffix = "%"

    return fig


def _fig_combo_overview(data: dict) -> go.Figure:
    """图2：融合曲线 + 全组合收益 & 回撤。"""
    combos = data["combos"]
    all_dates = data["all_dates"]
    assets = data["assets"]

    fig = make_subplots(
        rows=2, cols=1, shared_xaxes=True,
        row_heights=[0.65, 0.35], vertical_spacing=0.04,
        subplot_titles=["累计收益率 %", "回撤 %"],
    )

    # 各标的融合（细线）
    for i, a in enumerate(assets):
        k = f"{a}融合"
        if k not in combos:
            continue
        color = _asset_color(a, i)
        vals = combos[k]
        md = _max_dd(vals)
        fig.add_trace(go.Scatter(
            x=all_dates, y=vals,
            name=f"{k} (DD:{md:.1f}%)",
            line=dict(color=color, width=1.8, dash="dot"), opacity=0.7,
            legendgroup=k,
        ), row=1, col=1)
        fig.add_trace(go.Scatter(
            x=all_dates, y=_drawdown_series(vals),
            name=k, showlegend=False,
            line=dict(color=color, width=1, dash="dot"), opacity=0.5,
            legendgroup=k,
        ), row=2, col=1)

    # 全组合（粗线）
    import colorsys as _cs
    combo_keys = ["全组合", "全4标的", "全版本等权"] + [k for k in combos if "+" in k]
    non_fixed_colors = {}
    non_fixed = [k for k in combo_keys if k in combos and k not in ("全组合", "全4标的", "全版本等权")]
    for idx, k in enumerate(non_fixed):
        h = idx / max(len(non_fixed), 1)
        r2, g2, b2 = _cs.hsv_to_rgb(h, 0.75, 0.80)
        non_fixed_colors[k] = f'#{int(r2*255):02x}{int(g2*255):02x}{int(b2*255):02x}'

    for name in combo_keys:
        if name not in combos:
            continue
        vals = combos[name]
        md = _max_dd(vals)
        is_full = name in ("全组合", "全4标的", "全版本等权")
        color = "#E63946" if is_full else non_fixed_colors.get(name, "#888888")
        width = 3.5 if is_full else 2.2
        fig.add_trace(go.Scatter(
            x=all_dates, y=vals,
            name=f"{name} (DD:{md:.1f}%)",
            line=dict(color=color, width=width),
            legendgroup=name,
        ), row=1, col=1)
        fig.add_trace(go.Scatter(
            x=all_dates, y=_drawdown_series(vals),
            name=name, showlegend=False,
            line=dict(color=color, width=max(1, width - 1)),
            legendgroup=name,
        ), row=2, col=1)

    fig.update_layout(
        title="融合策略 & 组合收益 & 回撤",
        height=600, template="plotly_white", hovermode="x unified",
    )
    fig.update_yaxes(ticksuffix="%", row=1, col=1)
    fig.update_yaxes(ticksuffix="%", row=2, col=1)
    return fig


def _fig_corr_heatmap(data: dict) -> go.Figure:
    """图3：相关性热力图（月度 P&L）。"""
    monthly = data["monthly"]
    keys = list(monthly.keys())
    n = len(keys)

    # 对齐月份
    all_ym = set()
    for k in keys:
        all_ym |= set(monthly[k].keys())
    all_ym = sorted(all_ym)

    aligned = {k: [monthly[k].get(ym, 0.0) for ym in all_ym] for k in keys}

    matrix = []
    for ki in keys:
        row = []
        for kj in keys:
            r = _pearson(aligned[ki], aligned[kj])
            row.append(round(r, 3) if not math.isnan(r) else 0.0)
        matrix.append(row)

    text = [[f"{matrix[i][j]:.2f}" for j in range(n)] for i in range(n)]

    fig = go.Figure(go.Heatmap(
        z=matrix, x=keys, y=keys,
        text=text, texttemplate="%{text}", textfont=dict(size=10),
        colorscale=[[0, "#d73027"], [0.25, "#fc8d59"],
                    [0.5, "#ffffbf"], [0.75, "#91bfdb"], [1, "#4575b4"]],
        zmin=-1, zmax=1, zmid=0,
        colorbar=dict(title="r"),
        hovertemplate="%{y} × %{x}: %{text}<extra></extra>",
    ))
    fig.update_layout(
        title="策略月度收益相关性矩阵",
        height=max(400, n * 45 + 100),
        width=max(500, n * 60 + 150),
        template="plotly_white",
        xaxis=dict(tickangle=45),
    )
    return fig


def _fig_metrics_bar(data: dict, files: dict[str, str]) -> go.Figure:
    """图4：关键指标柱状对比（净利润、夏普、最大回撤、Sortino）。
    夏普/Sortino 用月度收益计算，与结论文档保持一致。
    """
    keys = list(files.keys())
    all_stats = data["raw_stats"]
    monthly = data["monthly"]
    series = data["series"]

    # 净利润和最大回撤从 xlsx 读，夏普/Sortino 用月度算（与结论文档一致）
    net_vals    = [_gk(all_stats[k], "净利润", "全部%") for k in keys]
    dd_vals     = [_gk(all_stats[k], "最大股权回撤（intrabar）", "全部%") for k in keys]
    sharpe_vals = [_sharpe(monthly.get(k, {})) for k in keys]
    sortino_vals= [_sortino(monthly.get(k, {})) for k in keys]

    metrics_data = [
        ("净利润 %",  net_vals,     True),
        ("夏普（月度）", sharpe_vals, False),
        ("最大回撤 %", dd_vals,     True),
        ("Sortino（月度）", sortino_vals, False),
    ]

    fig = make_subplots(
        rows=2, cols=2,
        subplot_titles=[m[0] for m in metrics_data],
        vertical_spacing=0.18, horizontal_spacing=0.1,
    )
    positions = [(1, 1), (1, 2), (2, 1), (2, 2)]

    for (label, y_vals, is_pct), (r, c) in zip(metrics_data, positions):
        colors = [_asset_color(k, i) for i, k in enumerate(keys)]
        fig.add_trace(go.Bar(
            x=keys, y=y_vals,
            marker_color=colors,
            showlegend=False,
            text=[f"{v:.2f}" for v in y_vals],
            textposition="outside",
        ), row=r, col=c)
        if is_pct:
            fig.update_yaxes(ticksuffix="%", row=r, col=c)

    fig.update_layout(
        title="各策略关键指标对比",
        height=600, template="plotly_white",
    )
    return fig


# ─── 结论文本生成 ─────────────────────────────────────────────────────────────

def _gen_conclusion(data: dict, files: dict[str, str]) -> str:
    """根据计算数据生成结论文本。"""
    combos = data["combos"]
    combo_monthly = data.get("combo_monthly", {})
    merged_monthly = data["merged_monthly"]
    monthly = data["monthly"]
    series = data["series"]
    assets = data["assets"]
    single_asset_mode = data["single_asset_mode"]
    raw_stats = data["raw_stats"]
    trades_data = data.get("trades", {})

    lines = []

    # ── 1. 各策略单独表现 ──────────────────────────────────────────────────────
    lines.append("## 各策略表现")
    lines.append("")
    lines.append("| 策略 | 净利润 % | 年化收益 % | 夏普 | Sortino | 最大回撤 % | 月胜率 % |")
    lines.append("|------|---------|-----------|------|---------|-----------|---------|")

    strat_stats = {}
    for k in sorted(files.keys()):
        pts = series.get(k, [])
        m = monthly.get(k, {})
        st = raw_stats.get(k, {})
        net = pts[-1] if pts else 0.0
        cagr = _gk(st, "年化收益率（CAGR）", "全部%")
        sharpe = _sharpe(m)
        sortino = _sortino(m)
        dd = _max_dd(pts) if pts else 0.0
        wr = sum(1 for v in m.values() if v > 0) / len(m) * 100 if m else 0.0
        strat_stats[k] = dict(net=net, cagr=cagr, sharpe=sharpe, sortino=sortino, dd=dd, wr=wr)
        lines.append(f"| {k} | {net:.1f}% | {cagr:.1f}% | {sharpe:.2f} | {sortino:.2f} | {dd:.1f}% | {wr:.0f}% |")

    lines.append("")

    # ── 2. 年度收益分解 ────────────────────────────────────────────────────────
    # 从交易清单按年汇总，转为占初始资本的百分比
    annual_by_strat: dict[str, dict[int, float]] = {}
    all_years: set[int] = set()
    for k in sorted(files.keys()):
        tlist = trades_data.get(k, [])
        apnl = _annual_pnl(tlist)
        annual_by_strat[k] = apnl
        all_years |= set(apnl.keys())

    if all_years:
        years_sorted = sorted(all_years)
        lines.append("## 年度收益分解")
        lines.append("")
        header_cols = " | ".join(str(y) for y in years_sorted)
        sep_cols = "|".join("------" for _ in years_sorted)
        lines.append(f"| 策略 | {header_cols} |")
        lines.append(f"|------|{sep_cols}|")
        for k in sorted(files.keys()):
            apnl = annual_by_strat.get(k, {})
            cells = []
            for y in years_sorted:
                v = apnl.get(y)
                if v is None:
                    cells.append("—")
                else:
                    pct = v / INITIAL * 100
                    cells.append(f"{pct:+.1f}%")
            lines.append(f"| {k} | {' | '.join(cells)} |")
        lines.append("")

    # ── 3. 交易统计 ────────────────────────────────────────────────────────────
    lines.append("## 交易统计")
    lines.append("")
    lines.append("| 策略 | 总交易数 | 胜率 % | 盈亏比 | 平均持仓K线 | 最大连续亏损 |")
    lines.append("|------|---------|-------|-------|-----------|------------|")
    for k in sorted(files.keys()):
        st = raw_stats.get(k, {})
        tlist = trades_data.get(k, [])
        # 总交易数：优先从 stats 读，fallback 从 trades 算
        total = int(_gk(st, "总交易", "全部USDT") or _gk(st, "Total trades", "全部USDT") or len(tlist))
        # 胜率
        wr_raw = (_gk(st, "获利百分比", "全部%") or _gk(st, "Percent profitable", "全部%"))
        win_rate = wr_raw if wr_raw > 1 else wr_raw * 100  # 有些是小数形式
        # 盈亏比
        ratio = (_gk(st, "平均胜率/平均负率", "全部USDT")
                 or _gk(st, "Ratio avg win / avg loss", "全部USDT"))
        # 平均持仓 K 线数
        avg_bars = int(_gk(st, "交易的平均#K线数", "全部USDT")
                       or _gk(st, "Avg # bars in trades", "全部USDT"))
        # 最大连续亏损
        max_loss_streak = _max_consec_loss(tlist)
        lines.append(f"| {k} | {total} | {win_rate:.1f}% | {ratio:.2f} | {avg_bars} | {max_loss_streak} |")
    lines.append("")

    # ── 4. 融合/全版本表现（多标的模式） ──────────────────────────────────────
    if not single_asset_mode:
        lines.append("## 各标的融合表现")
        lines.append("")
        lines.append("| 组合 | 净利润 % | 最大回撤 % | 夏普 | Sortino |")
        lines.append("|------|---------|-----------|------|---------|")
        for a in assets:
            k = f"{a}融合"
            if k not in combos:
                continue
            vals = combos[k]
            net = vals[-1] if vals else 0.0
            dd = _max_dd(vals)
            m = merged_monthly.get(a, {})
            lines.append(f"| {k} | {net:.1f}% | {dd:.1f}% | {_sharpe(m):.2f} | {_sortino(m):.2f} |")
        lines.append("")

    # ── 5. 组合对比（按最大回撤排序） ─────────────────────────────────────────
    lines.append("## 组合对比（按最大回撤排序）")
    lines.append("")
    lines.append("| 组合 | 净利润 % | 最大回撤 % | 夏普 | 回撤/收益 |")
    lines.append("|------|---------|-----------|------|---------|")

    combo_stats = []
    for name, vals in combos.items():
        if not vals:
            continue
        # 单标的模式跳过单个版本（只展示组合）
        if single_asset_mode and "+" not in name and name != "全版本等权":
            continue
        net = vals[-1]
        dd = _max_dd(vals)
        ratio = dd / net if net > 0.1 else float("inf")
        m = combo_monthly.get(name, {})
        sharpe = _sharpe(m)
        combo_stats.append((dd, name, net, ratio, sharpe))

    for dd, name, net, ratio, sharpe in sorted(combo_stats):
        ratio_str = f"{ratio:.2f}" if ratio != float("inf") else "N/A"
        sharpe_str = f"{sharpe:.2f}" if sharpe else "—"
        lines.append(f"| {name} | {net:.1f}% | {dd:.1f}% | {sharpe_str} | {ratio_str} |")

    lines.append("")

    # ── 4. 相关性摘要 ──────────────────────────────────────────────────────────
    lines.append("## 策略相关性")
    lines.append("")
    keys = list(files.keys())
    all_ym = set()
    for k in keys:
        all_ym |= set(monthly[k].keys())
    all_ym = sorted(all_ym)
    aligned = {k: [monthly[k].get(ym, 0.0) for ym in all_ym] for k in keys}

    corr_pairs = []
    for i in range(len(keys)):
        for j in range(i + 1, len(keys)):
            r = _pearson(aligned[keys[i]], aligned[keys[j]])
            if not math.isnan(r):
                corr_pairs.append((r, keys[i], keys[j]))

    high_corr = [(r, a, b) for r, a, b in corr_pairs if abs(r) > 0.6]
    low_corr  = sorted([(r, a, b) for r, a, b in corr_pairs if abs(r) <= 0.6],
                       key=lambda x: abs(x[0]))

    if high_corr:
        lines.append("**高相关策略对（|r| > 0.6，组合分散效果有限）：**")
        lines.append("")
        lines.append("| 策略 A | 策略 B | 相关系数 |")
        lines.append("|--------|--------|---------|")
        for r, a, b in sorted(high_corr, key=lambda x: -abs(x[0])):
            lines.append(f"| {a} | {b} | {r:.3f} |")
        lines.append("")

    if low_corr:
        lines.append("**低相关策略对（|r| ≤ 0.6，适合组合）：**")
        lines.append("")
        lines.append("| 策略 A | 策略 B | 相关系数 |")
        lines.append("|--------|--------|---------|")
        for r, a, b in low_corr[:8]:
            lines.append(f"| {a} | {b} | {r:.3f} |")
        lines.append("")

    # ── 5. 关键发现与建议 ──────────────────────────────────────────────────────
    lines.append("## 关键发现")
    lines.append("")

    # 最优单策略（夏普最高）
    best_sharpe_k = max(strat_stats, key=lambda k: strat_stats[k]["sharpe"])
    best_dd_k = min(strat_stats, key=lambda k: strat_stats[k]["dd"])
    lines.append(f"- **夏普最高**：{best_sharpe_k}（{strat_stats[best_sharpe_k]['sharpe']:.2f}）")
    lines.append(f"- **回撤最小**：{best_dd_k}（{strat_stats[best_dd_k]['dd']:.1f}%）")

    # 最优组合（回撤最小且净利润为正）
    valid_combos = [(dd, name, net) for dd, name, net, _, _s in combo_stats if net > 0]
    if valid_combos:
        best_combo = min(valid_combos, key=lambda x: x[0])
        lines.append(f"- **最优组合（回撤最小）**：{best_combo[1]}（回撤 {best_combo[0]:.1f}%，净利润 {best_combo[2]:.1f}%）")

    # 高相关警告
    if high_corr:
        pairs_str = "、".join(f"{a}×{b}" for _, a, b in sorted(high_corr, key=lambda x: -abs(x[0]))[:3])
        lines.append(f"- **注意**：{pairs_str} 相关性较高，同时持有分散效果有限")

    lines.append("")

    return "\n".join(lines)


# ─── 主入口 ───────────────────────────────────────────────────────────────────

def run(files: dict[str, str], base: str, out_dir: str,
        date_str: str | None = None, title: str = "策略分析"):
    """
    生成 Markdown 结论文档，内嵌 PNG 图表。

    files:    {key: 绝对路径}
    base:     路径前缀（通常为空字符串）
    out_dir:  输出目录
    date_str: 日期字符串
    title:    分析标题
    """
    if date_str is None:
        date_str = datetime.date.today().isoformat()

    # 加上 base 前缀
    abs_files = {k: base + v for k, v in files.items()}

    print("  计算数据...")
    data = _compute(abs_files)

    png_dir = Path(out_dir) / "png"
    png_dir.mkdir(exist_ok=True)

    # 生成并保存各图表 PNG
    print("  生成 PNG 图表...")
    figs = {
        "equity_overview": (_fig_equity_overview(data), "图1_收益曲线总览.png", 1200, 600),
        "combo_overview":  (_fig_combo_overview(data),  "图2_组合收益回撤.png", 1200, 600),
        "corr_heatmap":    (_fig_corr_heatmap(data),    "图3_相关性矩阵.png",   900,  600),
        "metrics_bar":     (_fig_metrics_bar(data, abs_files), "图4_关键指标对比.png", 1200, 600),
    }

    png_paths = {}
    for key, (fig, fname, w, h) in figs.items():
        png_path = str(png_dir / fname)
        _save_png(fig, png_path, width=w, height=h)
        png_paths[key] = f"png/{fname}"

    # 生成结论文本
    conclusion = _gen_conclusion(data, abs_files)

    # 组装 Markdown
    md_lines = [
        f"# {title} — 分析结论",
        "",
        f"生成时间：{date_str}",
        "",
        "---",
        "",
        "## 收益曲线总览",
        "",
        f"![收益曲线总览]({png_paths['equity_overview']})",
        "",
        "## 组合收益 & 回撤",
        "",
        f"![组合收益回撤]({png_paths['combo_overview']})",
        "",
        "## 关键指标对比",
        "",
        f"![关键指标对比]({png_paths['metrics_bar']})",
        "",
        "## 相关性矩阵",
        "",
        f"![相关性矩阵]({png_paths['corr_heatmap']})",
        "",
        "---",
        "",
        conclusion,
        "",
        "> 交互图表见同目录 HTML 文件。",
    ]

    md_path = Path(out_dir) / f"{title}_结论_{date_str}.md"
    md_path.write_text("\n".join(md_lines), encoding="utf-8")
    print(f"  结论文档已保存: {md_path}")
